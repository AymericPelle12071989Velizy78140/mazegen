from abc import ABC, abstractmethod
from pathlib import Path

import openpyxl

from grid.grid import Grid


class MazeBase(Grid, ABC):
    def __init__(self, maze_arg=None):
        # https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
        super().__init__()
        if type(maze_arg) is openpyxl.workbook.workbook.Workbook:
            self.build_from_workbook(maze_arg)
        elif type(maze_arg) is str:
            self.build_from_xslx(Path(maze_arg))
        elif type(maze_arg) is Path:
            self.build_from_xslx(maze_arg)
        else:
            raise ValueError(f"'{maze_arg}' has invalid type '{type(maze_arg)}'.")

    def build_from_xslx(self, xlsx_file: Path):
        wbook = openpyxl.load_workbook(str(xlsx_file))
        self.build_from_workbook(wbook)

    def build_from_workbook(self, wbook: openpyxl.workbook.workbook.Workbook):
        # Get sheets
        squares_sheet = wbook[self.__squares_sheet_name()]
        grounds_sheet = wbook[self.__grounds_sheet_name()]
        borders_sheet = wbook[self.__borders_sheet_name()]
        # Check sheets
        if squares_sheet is None:
            msg = f"Workbook '{wbook.path}' does not have the required sheet '{self.__squares_sheet_name()}'."
        if borders_sheet is None:
            msg = f"Workbook '{wbook.path}' does not have the required sheet '{self.__borders_sheet_name()}'."
        # Build squares from sheet cells
        mz_width = squares_sheet.max_column - squares_sheet.min_column + 1
        mz_height = squares_sheet.max_row - squares_sheet.min_row + 1
        self.build(mz_width, mz_height)
        j = 0
        for sj in range(squares_sheet.min_row, squares_sheet.max_row + 1):
            i = 0
            for si in range(squares_sheet.min_column, squares_sheet.max_column + 1):
                s_cell = squares_sheet.cell(row=sj, column=si)
                g_cell = grounds_sheet.cell(row=sj, column=si) if grounds_sheet is not None else None
                b_cell = borders_sheet.cell(row=sj, column=si)
                square = self._build_square_from_cells(s_cell, g_cell, b_cell)
                self.setitem(i, j, square)
                print(f"Maze[{i},{j}]: {square}")
                i += 1
            print()
            j += 1

    @staticmethod
    def __squares_sheet_name():
        return "squares"

    @staticmethod
    def __grounds_sheet_name():
        return "grounds"

    @staticmethod
    def __borders_sheet_name():
        return "borders"

    @abstractmethod
    def _build_square_from_cells(self, s_cell, g_cell, b_cell):
        pass
