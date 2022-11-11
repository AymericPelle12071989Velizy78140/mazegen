from pathlib import Path

import openpyxl
from parse import parse

import grid.grid
from direction4 import Direction4
from square import Square


class Maze(grid.grid.Grid):
    class Square(Square):
        def __init__(self):
            super().__init__()

        def __init__(self, cell: openpyxl.cell.cell.Cell):
            super().__init__()
            self.type = cell.value
            self.borders[Direction4.UP] = Maze.Square.__border_from_side(cell.border.top)
            self.borders[Direction4.LEFT] = Maze.Square.__border_from_side(cell.border.left)
            self.borders[Direction4.DOWN] = Maze.Square.__border_from_side(cell.border.bottom)
            self.borders[Direction4.RIGHT] = Maze.Square.__border_from_side(cell.border.right)

        @staticmethod
        def __border_from_side(side: openpyxl.styles.borders.Side):
            # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.borders.html?highlight=border#openpyxl.styles.borders.Border
            border = Maze.Square.Border()
            if side.style is not None:
                match side.style:
                    case "hair" | "thin" | "thick" | "medium":
                        border.type = "wall"
                    case _:
                        raise Exception("Border style not handled '{}'.".format(side.style))
                if side.color:
                    border.color = Maze.Square.__openpyxl_color_to_color_hexastr(side.color.rgb)
            return border

        @staticmethod
        def __openpyxl_color_to_color_hexastr(rgb):
            res = parse("{a:.2}{r:.2}{g:.2}{b:.2}", rgb)
            if res is None:
                return "black"
            print(rgb, res['r'], res['g'], res['b'], res['a'])
            return "#{}{}{}{}".format(res['r'], res['g'], res['b'], res['a'])
    # class

    def __init__(self, maze_arg=None):
        # https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
        if maze_arg is not None:
            if type(maze_arg) is openpyxl.workbook.workbook.Workbook:
                self.build_from_workbook(maze_arg)
            elif type(maze_arg) is str:
                self.build_from_xslx(maze_arg)
            elif type(maze_arg) is Path:
                self.build_from_xslx(str(maze_arg))
            else:
                raise ValueError("'{}' has invalid type '{}'.".format(maze_arg, type(maze_arg)))

    def build_from_xslx(self, xlsx_file: str):
        wbook = openpyxl.load_workbook(xlsx_file)
        self.build_from_workbook(wbook)

    def build_from_workbook(self, wbook: openpyxl.workbook.workbook.Workbook):
        mz_sheet = wbook["maze"]
        if mz_sheet:
            bg_sheet = wbook["background"] if "background" in wbook.sheetnames else None
            self.build_from_sheets(mz_sheet, bg_sheet)
        else:
            raise Exception("The provided workbook does not have a sheet named 'maze'")

    def build_from_sheets(self, mz_sheet: openpyxl.worksheet.worksheet.Worksheet,
                          bg_sheet: openpyxl.worksheet.worksheet.Worksheet):
        mz_width = mz_sheet.max_column - mz_sheet.min_column + 1
        mz_height = mz_sheet.max_row - mz_sheet.min_row + 1
        self.build(mz_width, mz_height)
        self.__fill_from_mz_sheet(mz_sheet)
        if bg_sheet:
            self.__fill_from_bg_sheet(bg_sheet, mz_sheet.min_column, mz_sheet.max_column,
                                      mz_sheet.min_row, mz_sheet.max_row)
        for j in range(self.height):
            for i in range(self.width):
                square = self.getitem(i, j)
                print("{},{}: {}".format(i, j, square))
                i += 1
            j += 1
            print()

    def __fill_from_mz_sheet(self, mz_sheet: openpyxl.worksheet.worksheet.Worksheet):
        j = 0
        for sj in range(mz_sheet.min_row, mz_sheet.max_row + 1):
            i = 0
            for si in range(mz_sheet.min_column, mz_sheet.max_column + 1):
                cell = mz_sheet.cell(row=sj, column=si)
                square = Maze.Square(cell)
                self.setitem(i, j, square)
                i += 1
            j += 1

    def __fill_from_bg_sheet(self, bg_sheet: openpyxl.worksheet.worksheet.Worksheet,
                             mz_minc, mz_maxc, mz_minr, mz_maxr):
        j = 0
        for sj in range(mz_minr, mz_maxr + 1):
            i = 0
            for si in range(mz_minc, mz_maxc + 1):
                cell = bg_sheet.cell(row=sj, column=si)
                square = self.getitem(i, j)
                square.background = cell.value
                i += 1
            j += 1
